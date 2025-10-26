import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from typing import List, Optional
from database import User, Profit
from db_operations import DatabaseOperations
import os


class ReportGenerator:
    
    @staticmethod
    async def generate_user_report(user_id: int, filename: Optional[str] = None) -> str:
        """Генерирует Excel отчет для пользователя"""
        user = await DatabaseOperations.get_user_by_id(user_id)
        if not user:
            raise ValueError("Пользователь не найден")
        
        # Получаем профиты
        profits = await DatabaseOperations.get_user_profits(user_id, limit=1000)
        stats = await DatabaseOperations.get_user_statistics(user_id)
        
        # Создаем DataFrame
        data = []
        for profit in profits:
            created_by = await DatabaseOperations.get_user_by_id(profit.created_by_id)
            data.append({
                'Дата': profit.created_at.strftime('%d.%m.%Y %H:%M'),
                'Сумма (₽)': profit.amount,
                'Добавил': created_by.username if created_by else 'Unknown',
                'Комментарий': profit.description or ''
            })
        
        df = pd.DataFrame(data)
        
        # Создаем Excel файл
        if not filename:
            filename = f"report_{user.username or user.telegram_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = os.path.join('reports', filename)
        os.makedirs('reports', exist_ok=True)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Лист с профитами
            df.to_excel(writer, sheet_name='Профиты', index=False)
            
            # Лист со статистикой
            stats_data = {
                'Показатель': [
                    'Общая сумма профитов',
                    'Количество профитов',
                    'За день',
                    'За неделю',
                    'За месяц',
                    'Максимальный занос',
                    'Средний занос'
                ],
                'Значение': [
                    f"{int(stats['total']):,} ₽",
                    stats['count'],
                    f"{int(stats['day']):,} ₽",
                    f"{int(stats['week']):,} ₽",
                    f"{int(stats['month']):,} ₽",
                    f"{int(stats['max_profit']):,} ₽",
                    f"{int(stats['avg_profit']):,} ₽"
                ]
            }
            stats_df = pd.DataFrame(stats_data)
            stats_df.to_excel(writer, sheet_name='Статистика', index=False)
            
            # Форматирование
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Заголовки
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Автоширина колонок
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return filepath
    
    @staticmethod
    async def generate_team_report(mini_head_id: int, filename: Optional[str] = None) -> str:
        """Генерирует Excel отчет для команды Mini Head"""
        mini_head = await DatabaseOperations.get_user_by_id(mini_head_id)
        if not mini_head:
            raise ValueError("Mini Head не найден")
        
        # Получаем команду
        team = await DatabaseOperations.get_mini_head_team(mini_head_id)
        team_stats = await DatabaseOperations.get_team_statistics(mini_head_id)
        
        # Создаем данные
        data = []
        for member in team:
            stats = await DatabaseOperations.get_user_statistics(member.id)
            status_data = member.get_status(stats['total'])
            data.append({
                'Telegram ID': member.telegram_id,
                'Username': member.username or 'не указан',
                'Статус': f"{status_data['emoji']} {status_data['name']}",
                'Общий профит (₽)': int(stats['total']),
                'Кол-во профитов': stats['count'],
                'За месяц (₽)': int(stats['month']),
                'Макс. занос (₽)': int(stats['max_profit']),
                'Средний занос (₽)': int(stats['avg_profit'])
            })
        
        df = pd.DataFrame(data)
        df = df.sort_values('Общий профит (₽)', ascending=False)
        
        # Создаем Excel файл
        if not filename:
            filename = f"team_report_{mini_head.username or mini_head.telegram_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = os.path.join('reports', filename)
        os.makedirs('reports', exist_ok=True)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Команда', index=False)
            
            # Общая статистика команды
            team_summary = {
                'Показатель': [
                    'Участников в команде',
                    'Общий профит команды',
                    'Количество профитов'
                ],
                'Значение': [
                    team_stats['members'],
                    f"{int(team_stats['total']):,} ₽",
                    team_stats['count']
                ]
            }
            summary_df = pd.DataFrame(team_summary)
            summary_df.to_excel(writer, sheet_name='Итоги', index=False)
            
            # Форматирование
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return filepath
    
    @staticmethod
    async def generate_global_report(filename: Optional[str] = None) -> str:
        """Генерирует глобальный отчет для Head"""
        # Получаем всех пользователей
        all_users = await DatabaseOperations.get_all_users()
        global_stats = await DatabaseOperations.get_global_statistics()
        
        # Создаем данные
        data = []
        for user in all_users:
            stats = await DatabaseOperations.get_user_statistics(user.id)
            status_data = user.get_status(stats['total'])
            
            mini_head_name = ''
            if user.mini_head_id:
                mini_head = await DatabaseOperations.get_user_by_id(user.mini_head_id)
                mini_head_name = mini_head.username if mini_head else ''
            
            data.append({
                'Telegram ID': user.telegram_id,
                'Username': user.username or 'не указан',
                'Роль': user.role,
                'Наставник': mini_head_name,
                'Статус': f"{status_data['emoji']} {status_data['name']}",
                'Общий профит (₽)': int(stats['total']),
                'Кол-во профитов': stats['count'],
                'За месяц (₽)': int(stats['month']),
                'FakeTag': user.fake_tag or ''
            })
        
        df = pd.DataFrame(data)
        df = df.sort_values('Общий профит (₽)', ascending=False)
        
        # Создаем Excel файл
        if not filename:
            filename = f"global_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = os.path.join('reports', filename)
        os.makedirs('reports', exist_ok=True)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Все пользователи', index=False)
            
            # Глобальная статистика
            global_summary = {
                'Показатель': [
                    'Всего пользователей',
                    'Активных пользователей',
                    'Общий профит',
                    'Количество профитов'
                ],
                'Значение': [
                    global_stats['total_users'],
                    global_stats['active_users'],
                    f"{int(global_stats['total']):,} ₽",
                    global_stats['count']
                ]
            }
            summary_df = pd.DataFrame(global_summary)
            summary_df.to_excel(writer, sheet_name='Глобальная статистика', index=False)
            
            # Форматирование
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return filepath
